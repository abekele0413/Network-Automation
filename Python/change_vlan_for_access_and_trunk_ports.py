#python3 /usr/local/bin/Access-port-script/accessvlanchange.py

#---------------------------------------------------------------------------------------------------------------------------------------------------


import netmiko
import time
from getpass import getpass
import os
import openpyxl
from openpyxl import Workbook
path = "/usr/local/bin/Access-port-script/devices.xlsx"
wb = openpyxl.load_workbook(path)
ws = wb.active
rowsvalue =ws.max_row
username = input('Username:')
password = getpass()

command3 = ''
command4 = ''
command5 = ''
command6 = ''
showrun = ''
ioscommandset = ''
nxcommandset = ''
command8 = ''
command9 = ''


command3 = "show run | i switchport access|interface"
command4 = 'show run vlan'
showrun = "show run"
command5 = "show run | i switchport trunk|interface"
#command6 = "show run | i tacacs"
#command7 = "show run | i tacacs"
#command8 = "show run | i V"

#ioscommandset = [
#'tacacs server VMware_SC2_ISE_Tacacs',

#]

#nxcommandset = [
#'tacacs server VMware_SC2_ISE_Tacacs',
#]




dictofvlans = {}
dictoftrunkvlans = {}
for i in range(2,rowsvalue+1):
	dtype = ws.cell(row = i, column = 2)
	dtype = dtype.value
	dtype = str(dtype)
	cell = ws.cell(row = i, column = 1)
	print('########################################################\nCurrently on row %s of %s. logging into %s\n########################################################'% (i,rowsvalue,cell.value))
	netdevice = cell.value
	listofvlans = []
	listoftrunkvlans = []
	if netdevice != None or '':
		try:
			if dtype == 'cisco_ios':
				net_connect = netmiko.ConnectHandler(device_type="cisco_ios", host=netdevice, username=username, password=password)
				prompt= net_connect.find_prompt()
				prompt=str(prompt)
				print(dictofvlans)
				print(dictoftrunkvlans)

						
				if command4 != '':
					output4 = net_connect.send_command(command4)
					ws.cell(row = i, column = 4, value = output4)
					print('Current Vlans on this switch:\n--------------------------------------------------\n'+output4)
				
				if command6 != '':
					output6 = net_connect.send_command(command6)
					ws.cell(row = i, column = 6, value = output6)	
			#-----------------conf t changes------------------------------------------------------------    

			#-----------------------------------------------------------------------------    
				if ioscommandset != '':
					output7 = net_connect.send_config_set(ioscommandset)
					ws.cell(row = i, column = 7, value = output7)
				if command8 != '':
					output8 = net_connect.send_command(command8)
					ws.cell(row = i, column = 8, value = output8)
				if command9 != '':
					output9 = net_connect.send_command(command9)
					ws.cell(row = i, column = 9, value = output9)
			#-----------------show run ------------------------------------------------------------    
			#-----------------------------------------------------------------------------    
				if showrun != '':
					showrunoutput = net_connect.send_command(showrun)
					showrunlength=len(showrunoutput)
					#print(showrunlength)
					showrunlengthloop = showrunlength // 25000
					#print(showrunlengthloop)
					valstart = 0
					valend = 25000
					colval2 = 10
					for l in range(showrunlengthloop+1):
						ws.cell(row = i, column = colval2, value = showrunoutput[valstart:valend])
						valstart += 25000
						valend += 25000
						colval2 +=1
					#columns 10-15

				if command3 != '':
					output3 = net_connect.send_command(command3)
					ws.cell(row = i, column = 3, value = output3)
					output3 = output3.splitlines()
					interfacevar=''
					oldvlan=''
					newvlan=''
					


					for vbs in output3:
						if 'switchport access vlan' in vbs and vbs not in listofvlans:
							listofvlans.append(vbs)
					print('-----------------------------------------------------------------\nACCESS PORTS\n-----------------------------------------------------------------')
					for bbb in range(0,len(listofvlans),1):
						#print(listofvlans[bbb])
						oldvlan32434543=listofvlans[bbb]
						if oldvlan32434543 in dictofvlans.keys():
							print('old vlan is already in the dictionary '+oldvlan32434543)
#						if oldvlan32434543 in dictofvlans.keys():
						else:	
							oldvlan32434543=oldvlan32434543[oldvlan32434543.find('switchport access vlan')+23:]
							##print(oldvlan32434543)
							vlname = net_connect.send_command('show run vlan %s | i name'%(oldvlan32434543))
							vlan23424=''
							vlan23424=input("\nHere is an old access port vlan "+oldvlan32434543+vlname+'\nWhat is the new vlan number or type "skip":')
							vlan23424=str(vlan23424)
							vlnameck = net_connect.send_command('show vlan id '+(vlan23424))
							while('not found in current VLAN database' in vlnameck or 'Command rejected' in vlnameck or 'Incomplete command' in vlnameck):
								if vlan23424=='skip' or vlan23424=='Skip':
									#print('\nvlan %s %s ports will be skipped'%(oldvlan32434543,vlname))
									break
								print('\nvlan not found, try again')
								vlan23424=input("Here is an old access port vlan "+oldvlan32434543+vlname+'\nWhat is the new vlan number:')
								vlan23424=str(vlan23424)
								vlnameck = net_connect.send_command('show vlan id '+(vlan23424))
							
						if vlan23424=='skip' or vlan23424=='Skip':
							print('\nvlan %s %s ports will be skipped'%(oldvlan32434543,vlname))
						#if vlan23424.startswith('0')or vlan23424.startswith('1')or vlan23424.startswith('2')or vlan23424.startswith('3')or vlan23424.startswith('4')or vlan23424.startswith('5')or vlan23424.startswith('6')or vlan23424.startswith('7')or vlan23424.startswith('8')or vlan23424.startswith('9'):
						else:
							dictofvlans[listofvlans[bbb]] = 'switchport access vlan '+vlan23424
							#vlnameck = net_connect.send_command('show vlan id '+(vlan23424))
							#while('not found in current VLAN database' in vlnameck or 'Command rejected' in vlnameck or vlan23424!='skip' or vlan23424!='Skip'):
							#	vlnameck = net_connect.send_command('show vlan id '+(vlan23424))
							#	print('\nvlan not found, try again')
							#	vlan23424=input("Here is an old vlan "+oldvlan32434543+vlname+'\nWhat is the new vlan number:')
							#	vlan23424=str(vlan23424)
								#if vlan23424=='skip' or vlan23424=='Skip':
								#	break

							
					


					
					
					commands1235=[]
					for vb in output3:
						vbbb = str(vb)
						if vbbb.startswith('interface')==True:

							interfacevar=vbbb
							#print(interfacevar)

						if 'switchport access vlan' in vbbb:
							#print(vbbb)
							for bbbbb in dictofvlans:
								#print(dictofvlans[bbbbb])
								#print(vbbb)
								#print(bbbbb)
								if vbbb == bbbbb:
									#print(interfacevar)
									#print('no '+vb)
									#print(dictofvlans[bbbbb]+'\n')
									commands1235.append(interfacevar)
									commands1235.append('no '+vb)
									commands1235.append(dictofvlans[bbbbb])
				if len(dictofvlans) ==0:
					print('No access ports found.')
					#print(commands1235)

					#listoftrunkvlans = []
					#dictoftrunkvlans = {}
				if command5 != '':
					output5 = net_connect.send_command(command5)
					ws.cell(row = i, column = 5, value = output5)
					output5 = output5.splitlines()
					interfacevar=''
					oldvlan=''
					newvlan=''
					
					print('-----------------------------------------------------------------\nTRUNK PORTS\n-----------------------------------------------------------------')
					

					for vbs in output5:
						if 'switchport trunk allowed' in vbs and vbs not in listoftrunkvlans:
							listoftrunkvlans.append(vbs)
					for bbb in range(0,len(listoftrunkvlans),1):
						#print(listoftrunkvlans[bbb])
						oldvlan32434543=listoftrunkvlans[bbb]
						if oldvlan32434543 in dictoftrunkvlans.keys():
							print('old vlan is already in the dictionary '+oldvlan32434543)
#						if oldvlan32434543 in dictofvlans.keys():
#							print('old vlan is already in the dictionary '+oldvlan32434543)
						oldvlan32434543=oldvlan32434543[oldvlan32434543.find('switchport trunk allowed')+30:]
						##print(oldvlan32434543)
						#vlname = net_connect.send_command('show run vlan %s | i name'%(oldvlan32434543))
						vlan23424=''
						vlan23424=input("\nHere are the old trunk vlans "+oldvlan32434543+'\nWhat are the new vlans for the trunk interfaces or type "skip". \n(multiple vlans should be formatted like this "100,101,102"):')
						vlan23424=str(vlan23424)
						vlnameck = net_connect.send_command('show vlan id '+(vlan23424))
						while('not found in current VLAN database' in vlnameck or 'Command rejected' in vlnameck or 'Incomplete command' in vlnameck):
							if vlan23424=='skip' or vlan23424=='Skip':
								#print('\nvlan %s %s ports will be skipped'%(oldvlan32434543,vlname))
								break
							print('\nvlan not found, try again')
							vlan23424=input("Here are the old trunk vlans "+oldvlan32434543+'\nWhat are the new vlans for the trunk interfaces or type "skip". \n(multiple vlans should be formatted like this "100,101,102"):')
							vlan23424=str(vlan23424)
							vlnameck = net_connect.send_command('show vlan id '+(vlan23424))
							
						if vlan23424=='skip' or vlan23424=='Skip':
							print('\nvlan %s ports will be skipped'%(oldvlan32434543))
						#if vlan23424.startswith('0')or vlan23424.startswith('1')or vlan23424.startswith('2')or vlan23424.startswith('3')or vlan23424.startswith('4')or vlan23424.startswith('5')or vlan23424.startswith('6')or vlan23424.startswith('7')or vlan23424.startswith('8')or vlan23424.startswith('9'):
						else:
							dictoftrunkvlans[listoftrunkvlans[bbb]] = 'switchport trunk allowed vlan add '+vlan23424
							#vlnameck = net_connect.send_command('show vlan id '+(vlan23424))
							#while('not found in current VLAN database' in vlnameck or 'Command rejected' in vlnameck or vlan23424!='skip' or vlan23424!='Skip'):
							#	vlnameck = net_connect.send_command('show vlan id '+(vlan23424))
							#	print('\nvlan not found, try again')
							#	vlan23424=input("Here is an old vlan "+oldvlan32434543+vlname+'\nWhat is the new vlan number:')
							#	vlan23424=str(vlan23424)
								#if vlan23424=='skip' or vlan23424=='Skip':
								#	break



					#print(dictoftrunkvlans) 
					
					for vb in output5:
						vbbb = str(vb)
						if vbbb.startswith('interface')==True:
							interfacevar=vbbb
							#print(interfacevar)

						if 'switchport trunk allowed' in vbbb:
							#print(vbbb)
							for bbbbb in dictoftrunkvlans:
								#print(dictoftrunkvlans[bbbbb])
								#print(vbbb)
								#print(bbbbb)
								if vbbb == bbbbb:
									portck = net_connect.send_command('show run '+(interfacevar))
									#print(portck)
									if 'channel-group' not in portck:
										#print(interfacevar)
										#print('no '+vb)
										#print(dictoftrunkvlans[bbbbb]+'\n')
										commands1235.append(interfacevar)
										commands1235.append(dictoftrunkvlans[bbbbb])
					#print(commands1235)
				print(dictofvlans)
				print(dictoftrunkvlans)
				print('\nPrepped Commands:')
				print('='*100)	
				for vbbbbbb in commands1235:
					print(vbbbbbb)
				print('='*100)
				input123=input("Would you like to run the commands?Y/N:")
				if input123 == 'Y' or input123 == 'y' or input123 == 'yes'or input123 == 'Yes'or input123 == 'YES':
					# to send commands!!!!!!!!!!!!!!!!!!!!!!
					#output7 = net_connect.send_config_set(commands1235)
					#ws.cell(row = i, column = 7, value = output7)
					print('commands ran')
				else:
					print('commands will not run')

			#-----------------------------------------------------------------------------
			#-----------------------------------------------------------------------------    	
			else:
				net_connect = netmiko.ConnectHandler(device_type="cisco_ios", host=netdevice, username=username, password=password)
				prompt= net_connect.find_prompt()
				prompt=str(prompt)
				print(dictofvlans)
				print(dictoftrunkvlans)
				if command4 != '':
					output4 = net_connect.send_command(command4)
					ws.cell(row = i, column = 4, value = output4)
					print('Current Vlans on this switch:\n--------------------------------------------------\n'+output4)
				
				if command6 != '':
					output6 = net_connect.send_command(command6)
					ws.cell(row = i, column = 6, value = output6)	
			#---	--------------conf t changes------------------------------------------------------------    
			#---	--------------------------------------------------------------------------    
				if nxcommandset != '':
					output5 = net_connect.send_config_set(nxcommandset)
					ws.cell(row = i, column = 7, value = output7)
				if command8 != '':
					output8 = net_connect.send_command(command8)
					ws.cell(row = i, column = 8, value = output8)
				if command9 != '':
					output9 = net_connect.send_command(command9)
					ws.cell(row = i, column = 9, value = output9)
		
			#--	--------------show run ------------------------------------------------------------    
			#---	--------------------------------------------------------------------------    
				if showrun != '':
					showrunoutput = net_connect.send_command(showrun)
					showrunlength=len(showrunoutput)
					#print(showrunlength)
					showrunlengthloop = showrunlength // 25000
					#print(showrunlengthloop)
					valstart = 0
					valend = 25000
					colval2 = 10
					for l in range(showrunlengthloop+1):
						ws.cell(row = i, column = colval2, value = showrunoutput[valstart:valend])
						valstart += 25000
						valend += 25000
						colval2 +=1
					#columns 10-15
			#------------------------------------------------------------------------------
				if command3 != '':
					output3 = net_connect.send_command("show run | i 'switchport access|interface'")
					ws.cell(row = i, column = 3, value = output3)
					output3 = output3.splitlines()
					interfacevar=''
					oldvlan=''
					newvlan=''
					
				# contuniue  exsisting list 			
				#if len(dictofvlans) == 0:

					for vbs in output3:
						if 'switchport access vlan' in vbs and vbs not in listofvlans:
							listofvlans.append(vbs)
					print('-----------------------------------------------------------------\nACCESS PORTS\n-----------------------------------------------------------------')
					
					for bbb in range(0,len(listofvlans),1):
						#print(listofvlans[bbb])
						oldvlan32434543=listofvlans[bbb]
						if oldvlan32434543 in dictofvlans.keys():
							print('old vlan is already in the dictionary '+oldvlan32434543)
#						if oldvlan32434543 in dictofvlans.keys():
						else:	
							oldvlan32434543=oldvlan32434543[oldvlan32434543.find('switchport access vlan')+23:]
							##print(oldvlan32434543)
							vlname = net_connect.send_command('show run vlan %s | i name'%(oldvlan32434543))
							vlan23424=''
							vlan23424=input("\nHere is an old access port vlan "+oldvlan32434543+vlname+'\nWhat is the new vlan number or type "skip":')
							vlan23424=str(vlan23424)
							vlnameck = net_connect.send_command('show vlan id '+(vlan23424))
							while('not found in current VLAN database' in vlnameck or 'Command rejected' in vlnameck or 'Incomplete command' in vlnameck):
								if vlan23424=='skip' or vlan23424=='Skip':
									#print('\nvlan %s %s ports will be skipped'%(oldvlan32434543,vlname))
									break
								print('\nvlan not found, try again')
								vlan23424=input("Here is an old access port vlan "+oldvlan32434543+vlname+'\nWhat is the new vlan number:')
								vlan23424=str(vlan23424)
								vlnameck = net_connect.send_command('show vlan id '+(vlan23424))
							
						if vlan23424=='skip' or vlan23424=='Skip':
							print('\nvlan %s %s ports will be skipped'%(oldvlan32434543,vlname))
						#if vlan23424.startswith('0')or vlan23424.startswith('1')or vlan23424.startswith('2')or vlan23424.startswith('3')or vlan23424.startswith('4')or vlan23424.startswith('5')or vlan23424.startswith('6')or vlan23424.startswith('7')or vlan23424.startswith('8')or vlan23424.startswith('9'):
						else:
							dictofvlans[listofvlans[bbb]] = 'switchport access vlan '+vlan23424
							#vlnameck = net_connect.send_command('show vlan id '+(vlan23424))
							#while('not found in current VLAN database' in vlnameck or 'Command rejected' in vlnameck or vlan23424!='skip' or vlan23424!='Skip'):
							#	vlnameck = net_connect.send_command('show vlan id '+(vlan23424))
							#	print('\nvlan not found, try again')
							#	vlan23424=input("Here is an old vlan "+oldvlan32434543+vlname+'\nWhat is the new vlan number:')
							#	vlan23424=str(vlan23424)
								#if vlan23424=='skip' or vlan23424=='Skip':
								#	break

							
					


					#print(dictofvlans) 

					commands1235=[]
					for vb in output3:
						vbbb = str(vb)
						if vbbb.startswith('interface')==True:
							interfacevar=vbbb
							#print(interfacevar)

						if 'switchport access vlan' in vbbb:
							#print(vbbb)
							for bbbbb in dictofvlans:
								#print(dictofvlans[bbbbb])
								#print(vbbb)
								#print(bbbbb)
								if vbbb == bbbbb:
									#print(interfacevar)
									#print('no '+vb)
									#print(dictofvlans[bbbbb]+'\n')
									commands1235.append(interfacevar)
									commands1235.append('no '+vb)
									commands1235.append(dictofvlans[bbbbb])
				if len(dictofvlans) ==0:
					print('No access ports found.')

					#print(commands1235)
				if command5 != '':
					output5 = net_connect.send_command("show run | i 'switchport trunk|interface'")
					ws.cell(row = i, column = 5, value = output5)
					output5 = output5.splitlines()
					interfacevar=''
					oldvlan=''
					newvlan=''
					
					print('-----------------------------------------------------------------\nTRUNK PORTS\n-----------------------------------------------------------------')
					

					for vbs in output5:
						if 'switchport trunk allowed' in vbs and vbs not in listoftrunkvlans:
							listoftrunkvlans.append(vbs)
					for bbb in range(0,len(listoftrunkvlans),1):
						#print(listoftrunkvlans[bbb])
						oldvlan32434543=listoftrunkvlans[bbb]
#						if oldvlan32434543 in dictofvlans.keys():
#							print('old vlan is already in the dictionary '+oldvlan32434543)
						oldvlan32434543=oldvlan32434543[oldvlan32434543.find('switchport trunk allowed')+30:]
						#print(oldvlan32434543)
						#vlname = net_connect.send_command('show run vlan %s | i name'%(oldvlan32434543))
						vlan23424=''
						vlan23424=input("\nHere are the old trunk vlans "+oldvlan32434543+'\nWhat are the new vlans for the trunk interfaces or type "skip". \n(multiple vlans should be formatted like this "100,101,102"):')
						vlan23424=str(vlan23424)
						vlnameck = net_connect.send_command('show vlan id '+(vlan23424))
						while('not found in current VLAN database' in vlnameck or 'Command rejected' in vlnameck or 'Incomplete command' in vlnameck):
							if vlan23424=='skip' or vlan23424=='Skip':
								#print('\nvlan %s %s ports will be skipped'%(oldvlan32434543,vlname))
								break
							print('\nvlan not found, try again')
							vlan23424=input("Here are the old trunk vlans "+oldvlan32434543+'\nWhat are the new vlans for the trunk interfaces or type "skip". \n(multiple vlans should be formatted like this "100,101,102"):')
							vlan23424=str(vlan23424)
							vlnameck = net_connect.send_command('show vlan id '+(vlan23424))
							
						if vlan23424=='skip' or vlan23424=='Skip':
							print('\nvlan %s ports will be skipped'%(oldvlan32434543))
						#if vlan23424.startswith('0')or vlan23424.startswith('1')or vlan23424.startswith('2')or vlan23424.startswith('3')or vlan23424.startswith('4')or vlan23424.startswith('5')or vlan23424.startswith('6')or vlan23424.startswith('7')or vlan23424.startswith('8')or vlan23424.startswith('9'):
						else:
							dictoftrunkvlans[listoftrunkvlans[bbb]] = 'switchport trunk allowed vlan add '+vlan23424
							#vlnameck = net_connect.send_command('show vlan id '+(vlan23424))
							#while('not found in current VLAN database' in vlnameck or 'Command rejected' in vlnameck or vlan23424!='skip' or vlan23424!='Skip'):
							#	vlnameck = net_connect.send_command('show vlan id '+(vlan23424))
							#	print('\nvlan not found, try again')
							#	vlan23424=input("Here is an old vlan "+oldvlan32434543+vlname+'\nWhat is the new vlan number:')
							#	vlan23424=str(vlan23424)
								#if vlan23424=='skip' or vlan23424=='Skip':
								#	break



					
					
					for vb in output5:
						vbbb = str(vb)
						if vbbb.startswith('interface')==True:
							interfacevar=vbbb
							#print(interfacevar)

						if 'switchport trunk allowed' in vbbb:
							#print(vbbb)
							for bbbbb in dictoftrunkvlans:
								#print(dictoftrunkvlans[bbbbb])
								#print(vbbb)
								#print(bbbbb)
								if vbbb == bbbbb:
									portck = net_connect.send_command('show run '+(interfacevar))
									#print(portck)
									if 'channel-group' not in portck:
										#print(interfacevar)
										#print('no '+vb)
										#print(dictoftrunkvlans[bbbbb]+'\n')
										commands1235.append(interfacevar)
										commands1235.append(dictoftrunkvlans[bbbbb])
					#print(commands1235)
				print(dictofvlans)
				print(dictoftrunkvlans)
				print('\nPrepped Commands:')
				print('='*100)	
				for vbbbbbb in commands1235:
					print(vbbbbbb)
				print('='*100)
				input123=input("Would you like to run the commands?Y/N:")
				if input123 == 'Y' or input123 == 'y' or input123 == 'yes'or input123 == 'Yes'or input123 == 'YES':
					# to send commands!!!!!!!!!!!!!!!!!!!!!!
					#output7 = net_connect.send_config_set(commands1235)
					#ws.cell(row = i, column = 7, value = output7)
					print('commands ran')
				else:
					print('commands will not run')


			net_connect.disconnect()
			wb.save(path)
			#print('########################################################\nCurrently on row %s of %s. Logged into %s\n---------------------------------------'% (i,rowsvalue,cell.value))
		#			print('======================================================== \n %s before: \n %s \n------------------------------------\n %s after: \n %s\n ========================================================'%(netdevice,output2,netdevice, output6))
				
		#			x=input('%s was sucessful. press enter to continue '%(netdevice))
		except:
			ws.cell(row = i, column = 3, value = 'there was an error here')
			wb.save(path)
			print('error connecting')
#			x=input('%s had issues logging in. press enter to continue'%(netdevice))
#print('all commands ran :)')



# contuniue  exsisting list 			
#					if len(dictofvlans) == 0:
