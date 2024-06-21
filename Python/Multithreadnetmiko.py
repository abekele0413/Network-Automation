#python3 /mnt/c/Users/abekele/'OneDrive - VMware, Inc'/Documents/sshnetmikomultithread.py

#---------------------------------------------------------------------------------------------------------------------------------------------------
#python3 /mnt/c/Users/abekele/'OneDrive - VMware, Inc'/Documents/sshnetmikomultithread.py

import netmiko
import time
from getpass import getpass
import os
import openpyxl
from openpyxl import Workbook
import threading
import json

#path = "/mnt/c/Users/abekele/Documents/devices.xlsx"
path = "/home/abekele/devicesmultithread.xlsx"
wb = openpyxl.load_workbook(path)
ws = wb.active
rowsvalue =ws.max_row

username = "svc.isebackup"
passwdgetpass = getpass()
password = passwdgetpass+"4"

command3 = ''
command4 = ''
command5 = ''
command6 = ''
showrun = ''
ioscommandset = ''
nxcommandset = ''
command8 = ''
command9 = ''
output3 = ''
output4 = ''
output5 = ''
output6 = ''
output7 = ''
output8 = ''
output9 = ''
rowcount2 = 0
rowcount3 = 0
rowcount4 = 0
rowcount5 = 0
rowcount6 = 0
rowcount7 = 0
rowcount8 = 0
rowcount9 = 0
rowcount10 = 0
rowcount11 = 0
rowcount12 = 0
rowcount13 = 0
rowcount14 = 0
rowcount15 = 0
rowcount16 = 0
#-------------------show commands------------------------------------------------------------

command3 = "show ver | i uptime"
#command4 = "show run | sec aaa"
#command8 = 'show vlan br'
showrun = "show run"
#command4 = "show run | i vrf"

#-------------------show commands------------------------------------------------------------


#-------------------conf t changes------------------------------------------------------------

#ioscommandset = [
#"no aaa authentication login default group ISE-T-ADMIN_LAB04 local",
#"no aaa authentication login AUTHEN-ADMIN-VTY-R group ISE-T-ADMIN_LAB04 local-case",
#"no aaa authorization exec AUTHOR-EXEC-VTY-R group ISE-T-ADMIN_LAB04 local ",
#"no aaa authorization commands 15 default group ISE-T-ADMIN_LAB04 local ",
#"aaa authentication login default group ISE-R-ADMIN_LAB04 local",
#"aaa authentication login AUTHEN-ADMIN-VTY-R group ISE-R-ADMIN_LAB04 local-case",
#"aaa authorization exec AUTHOR-EXEC-VTY-R group ISE-R-ADMIN_LAB04 local ",
#]

#nxcommandset = [
#''
#]
#-------------------conf t changes------------------------------------------------------------


commandlista = [command3,command3,command3,command3, command4, command5, command6, ioscommandset, command8, command9]
commandlistb = [command3,command3,command3,command3, command4, command5, command6, nxcommandset, command8, command9]
outputlist =   [output3, output3, output3, output3,  output4,  output5,  output6,  output7,      output8,  output9]
rowcountlist= [rowcount2,rowcount2,rowcount2,rowcount3,rowcount4,rowcount5,rowcount6,rowcount7,rowcount8,rowcount9,rowcount10,rowcount11,rowcount12,rowcount13,rowcount14,rowcount15,rowcount16]

print('running....')
def netmikoscript(rowcount):
	rowcountlist[rowcount]=rowcount
	for i in range(2,rowsvalue+1):
		if i ==rowcountlist[rowcount]:
			dtype = ws.cell(row = i, column = 2)
			dtype = dtype.value
			dtype = str(dtype)
			cell = ws.cell(row = i, column = 1)
			netdevice = cell.value
			if netdevice != None or '':
				try:
					if dtype == 'cisco_ios':
						net_connect = netmiko.ConnectHandler(device_type="cisco_ios", host=netdevice, username=username, password=password,global_delay_factor=2)
						prompt= net_connect.find_prompt()
						prompt=str(prompt)
						for vvv in range(3,10):
							if commandlista[vvv] != '':
								if commandlista[vvv] == ioscommandset:
									outputlist[vvv] = net_connect.send_config_set(commandlista[vvv])
									ws.cell(row = i, column = vvv, value = outputlist[vvv])
								else:
									outputlist[vvv]= net_connect.send_command(commandlista[vvv])
									ws.cell(row = i, column = vvv, value = outputlist[vvv])
			#-------------------show run ------------------------------------------------------------    
			#-------------------------------------------------------------------------------    
						if showrun != '':
							showrunoutput = net_connect.send_command(showrun)
							showrunlength=len(showrunoutput)
							#print(showrunlength)
							showrunlengthloop = showrunlength // 25000
							#print(showrunlengthloop)
							valstart = 0
							valend = 25000
							colval2 = 10
							for l in range(showrunlengthloop+1):
								ws.cell(row = i, column = colval2, value = showrunoutput[valstart:valend])
								valstart += 25000
								valend += 25000
								colval2 +=1
							#columns 10-15
			#-------------------------------------------------------------------------------
			#-------------------------------------------------------------------------------    
					else:
						net_connect = netmiko.ConnectHandler(device_type="cisco_ios", host=netdevice, username=username, password=password)
						prompt= net_connect.find_prompt()
						prompt=str(prompt)
						for vvv in range(3,10):
							if commandlistb[vvv] != '':
								if commandlistb[vvv] == nxcommandset:
									outputlist[vvv] = net_connect.send_config_set(commandlistb[vvv])
									ws.cell(row = i, column = vvv, value = outputlist[vvv])
								else:
									outputlist[vvv]= net_connect.send_command(commandlistb[vvv])
									ws.cell(row = i, column = vvv, value = outputlist[vvv])
			
			#-------------------show run ------------------------------------------------------------    
			#-------------------------------------------------------------------------------    
						if showrun != '':
							showrunoutput = net_connect.send_command(showrun)
							showrunlength=len(showrunoutput)
							#print(showrunlength)
							showrunlengthloop = showrunlength // 25000
							#print(showrunlengthloop)
							valstart = 0
							valend = 25000
							colval2 = 10
							for l in range(showrunlengthloop+1):
								ws.cell(row = i, column = colval2, value = showrunoutput[valstart:valend])
								valstart += 25000
								valend += 25000
								colval2 +=1
							#columns 10-15
			#-------------------------------------------------------------------------------

					net_connect.disconnect()
					wb.save(path)
					print('########################################################\nCurrently on row %s of %s. Logged into %s\n---------------------------------------'% (i,rowsvalue,cell.value))
		#			print('======================================================== \n %s before: \n %s \n------------------------------------\n %s after: \n %s\n ========================================================'%(netdevice,output2,netdevice, output6))
				
		#			x=input('%s was sucessful. press enter to continue '%(netdevice))
				except:
					print('########################################################\nCurrently on row %s of %s. Error logging into %s\n---------------------------------------'% (i,rowsvalue,cell.value))
					ws.cell(row = i, column = 3, value = 'there was an error here')
					wb.save(path)
					print('error connecting')
		#			x=input('%s had issues logging in. press enter to continue'%(netdevice))
			rowcountlist[rowcount]+=10





def MTstartrow2():
	netmikoscript(2)
def MTstartrow3():
	netmikoscript(3)
def MTstartrow4():
	netmikoscript(4)
def MTstartrow5():
	netmikoscript(5)
def MTstartrow6():
	netmikoscript(6)
def MTstartrow7():
	netmikoscript(7)
def MTstartrow8():
	netmikoscript(8)
def MTstartrow9():
	netmikoscript(9)
def MTstartrow10():
	netmikoscript(10)
def MTstartrow11():
	netmikoscript(11)
def MTstartrow12():
	netmikoscript(12)
def MTstartrow13():
	netmikoscript(13)
def MTstartrow14():
	netmikoscript(14)
def MTstartrow15():
	netmikoscript(15)
def MTstartrow16():
	netmikoscript(16)

threading.Thread(target=MTstartrow2).start()
threading.Thread(target=MTstartrow3).start()
threading.Thread(target=MTstartrow4).start()
threading.Thread(target=MTstartrow5).start()
threading.Thread(target=MTstartrow6).start()
threading.Thread(target=MTstartrow7).start()
threading.Thread(target=MTstartrow8).start()
threading.Thread(target=MTstartrow9).start()
threading.Thread(target=MTstartrow10).start()
threading.Thread(target=MTstartrow11).start()
threading.Thread(target=MTstartrow12).start()
threading.Thread(target=MTstartrow13).start()
threading.Thread(target=MTstartrow14).start()
threading.Thread(target=MTstartrow15).start()
threading.Thread(target=MTstartrow16).start()


#---------------------------------------------------------------
#corralating data with textfsm and adding it to dictionary
#output3 = net_connect.send_command('show vlan br',use_textfsm=True)
#output4 = net_connect.send_command('show ip arp',use_textfsm=True)
#
#for i in output3:
#	ii=('Vlan'+i['vlan_id'])
#	for l in output4:
#		if str(l['interface']) == str(ii):
#			newdata={'vlan_name':i['name']}
#			l.update(newdata)
#
#for i in output4:
#	print(i['interface']+' '+i['vlan_name'])

#				if command4 != '':
#					output4 = net_connect.send_command(command4)
#					ws.cell(row = i, column = 4, value = output4)
#					print(output4)
#				if command5 != '':
#					output5 = net_connect.send_command(command5)
#					ws.cell(row = i, column = 5, value = output5)
#				if command6 != '':
#					output6 = net_connect.send_command(command6)
#					ws.cell(row = i, column = 6, value = output6)	
#			#---	--------------conf t changes------------------------------------------------------------    
#			#---	--------------------------------------------------------------------------    
#				if nxcommandset != '':
#					output5 = net_connect.send_config_set(nxcommandset)
#					ws.cell(row = i, column = 7, value = output7)
#				if command8 != '':
#					output8 = net_connect.send_command(command8)
#					ws.cell(row = i, column = 8, value = output8)
#				if command9 != '':
#					output9 = net_connect.send_command(command9)
#					ws.cell(row = i, column = 9, value = output9)
